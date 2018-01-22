#!/usr/bin/env python
import urllib2
from openpyxl import *
from openpyxl.styles import *
from openpyxl.chart.shapes import GraphicalProperties
import csv, time,datetime,re
from nsetools import Nse
from pprint import pprint
import locale, operator
from openpyxl.chart import (
    BarChart3D,
    Reference,
    Series,
)


nifty50list = []
nift50dic = {}


def get_nifty50_list():
    global nifty50list
    try:
        wb = os.path.dirname(os.path.abspath(__file__))+"\\nifty50list.csv"   
        with open(wb, 'rb') as f:
            reader = csv.reader(f)
            for row in reader:
                if row[2].lower() != "symbol":
                    nifty50list.append(row[2])            
    except Exception, e:
        print e
       


def add_comma(value):
    return value
    try:       
        locale.setlocale(locale.LC_NUMERIC, "")
        val = locale.format("%.2f", value, grouping=True)        
        return val 
    except Exception,e:
        return value



def date(date):
    try:
        d1 = re.findall(r'\d+',date)
        d2 = re.findall(r'\D+',date)
        d = d1[0]+"-"+d2[0]+"-"+d1[1]    
        return d    
    except Exception,e:
        return date


def curr_date():
    try:
        now = datetime.datetime.now()
        now = str(now)
        return now.split(' ')[0]
    except Exception,e:
            return "None"

def time():
    try:
        now = datetime.datetime.now()
        now = str(now)
        return now.split(' ')[1].split('.')[0]
    except Exception,e:
        return "None"


def chart(ws):    
    row_count = ws.max_row    
    data = Reference(ws, min_col=4, max_col=5, min_row=3, max_row=row_count)
    titles = Reference(ws, min_col=1, min_row=4, max_row=row_count)
    chart = BarChart3D()    
    #chart.plot_area.graphicalProperties = GraphicalProperties(solidFill="000000")
    chart.x_axis.title = "Date"
    chart.y_axis.title = "Price"
    chart.title = "3D Bar Chart"
    chart.add_data(data=data, titles_from_data=True)
    chart.set_categories(titles)
    ws.add_chart(chart, "Q4")



def NSE():
    dic = { 'A3':'Date', 'B3':'Time', 'C3': 'Prev Close', 'D3':'Open',
            'E3':'Close', 'F3':'High', 'G3':'Low', 'H3':'LTP', 'I3':'Avg.Price',
            'J3':'INR Change','K3':'% Change', 'L3':'Volume', 'M3':'Turnover (in Lakhs)',
            'N3':'52-WK High', 'O3':'52-WK Low'
            }
    get_nifty50_list()
    try:
        wb = os.path.dirname(os.path.abspath(__file__))+"\\Data\\Nifty50_Stocks_History_Data.xlsx"    
        if not os.path.exists(wb):
            workbook = Workbook()        
            workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))
        else:        
            workbook = load_workbook(wb)
            
        sheets = workbook.get_sheet_names()                   
        nse = Nse()
        all_stock_codes = nse.get_stock_codes(cached=False)        
        for stock in sorted(all_stock_codes.iterkeys()):
            if stock in nifty50list:
                try:                   
                    st = nse.get_quote(stock)                    
                except urllib2.HTTPError, e:
                    print e.code
                    return 0
                except urllib2.URLError, e:
                    print e.args
                    return 0
                if st['change'] == None:
                    inr_change = st['lastPrice'] - st['previousClose']
                else:
                    inr_change = st['change']
                    
                if st['pChange'] == None:
                    per_change =  "{0:.2f}".format(((st['lastPrice'] - st['previousClose'])/st['previousClose'])*100)
                else:
                    per_change = st['pChange']
                    
                nift50dic[stock] = [curr_date(),time(),add_comma(st['previousClose']),add_comma(st['open']),add_comma(st['closePrice']),\
                                    add_comma(st['dayHigh']),add_comma(st['dayLow']),add_comma(st['lastPrice']),add_comma(st['averagePrice']),\
                                    inr_change, per_change, add_comma(st['totalTradedVolume']),add_comma(st['totalTradedValue']),\
                                    add_comma(st['high52']),add_comma(st['low52'])]
                if stock not in sheets:
                    worksheet = workbook.create_sheet(stock)
                    worksheet.freeze_panes = 'A4'
                    worksheet.merge_cells('A1:O2')
                    cell = worksheet.cell(row=1, column=1)
                    cell.value = stock + ":    " + st['companyName']
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(color=colors.BLUE, italic=True, name='Arial', size=14, bold=True, underline="single")
                    cell.fill = PatternFill(start_color='FFDEAD', end_color='FFDEAD',fill_type='solid')
                    for item in dic:
                        worksheet[item].alignment =  worksheet[item].alignment.copy(wrapText=True)
                        worksheet[item] = dic[item]
                        worksheet[item].font = Font(color=colors.RED, name='Arial', size=10, bold=True,italic=True)
                        worksheet[item].fill = PatternFill(start_color='000000', end_color='000000',fill_type='solid')               

                                        
                else:
                    worksheet = workbook.get_sheet_by_name(stock)
                worksheet.append(nift50dic[stock])
                chart(worksheet)
                break
                                                                                                                
                       
        workbook.save(wb)
        return 1
        
    except Exception,e:
        print e
        return 0


#for key, value in sorted(mydict.iteritems(), key=lambda (k,v): (v,k)):
        
            
if __name__ == "__main__":
    try:
        print "Getting Data...\n"        
        if NSE():
            print "Data store is done!!!"
    except Exception,e:
        print e

        
    

